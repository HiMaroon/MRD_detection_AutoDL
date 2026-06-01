import os
import re
import json
import collections
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


# =========================================================
# 1. 配置区
# =========================================================

DIRECTORIES = [
    "/root/autodl-tmp/data/MAIN_imgs_260323",
    "/root/autodl-tmp/data/TJMU_imgs_260416",
    "/root/autodl-tmp/data/BJH_imgs_260211",
    "/root/autodl-tmp/data/FXH_imgs_noALL_260318",
]

PATIENT_INFO_XLSX = "/root/autodl-tmp/data/patient_data_260416.xlsx"
PATIENT_INFO_SHEET = "总表"

OUTPUT_DIR = "/root/autodl-tmp/data/patient_NM_ratio_distribution_260416"

# 是否只统计 polygon 标注
POLYGON_ONLY = True

# 患者 ID 提取方式：
# first_level: 每个主目录下第一层子文件夹为患者，例如 MAIN_imgs_260323/PKUPH-001/xxx.json
# full_relative: 使用完整相对路径作为患者 ID
PATIENT_ID_MODE = "first_level"

# N / M 细胞比例定义
# 默认只统计精确标签 "N" 和 "M"
# 如需合并更多标签，可以改成：
# N_LABELS = ["N0", "N", "N1", "N2", "N3", "N4", "N5"]
# M_LABELS = ["M0", "M", "M1", "M2"]
N_LABELS = ["N"]
M_LABELS = ["M"]


# =========================================================
# 2. 作图参数
# =========================================================

FIG_DPI = 350
FIG_WIDTH = 9.5
FIG_HEIGHT = 7.8

# 散点大小
POINT_SIZE = 24

# 散点透明度
POINT_ALPHA = 0.85

# 是否标注 5% 参考线
SHOW_REF_LINE = True
REF_RATIO = 5.0

# 是否标注患者名称；点多时建议 False
ANNOTATE_PATIENT_ID = False
ANNOTATE_FONTSIZE = 6


# =========================================================
# 3. 显示名称映射
# =========================================================

CENTER_DISPLAY_MAP = {
    "PKUPH": "北大人民医院 PKUPH",
    "北大人民医院": "北大人民医院 PKUPH",
    "北京大学人民医院": "北大人民医院 PKUPH",

    "TAB": "荻硕贝肯 TAB",
    "荻硕贝肯": "荻硕贝肯 TAB",

    "BEPH": "北京电力医院 BEPH",
    "北京电力医院": "北京电力医院 BEPH",

    "FXH": "复兴医院 FXH",
    "首都医科大学附属复兴医院": "复兴医院 FXH",

    "BJH": "北京医院 BJH",
    "北京医院血液实验室": "北京医院血液实验室 BJH",

    "TJMU": "华中科技大学同济医学院 TJMU",
    "华中科技大学同济医学院": "华中科技大学同济医学院 TJMU",

    "MAIN": "MAIN",
}

PATIENT_TYPE_DISPLAY_MAP = {
    "AML": "AML患者",
    "HC": "非AML患者（正常人）",
    "NORMAL": "非AML患者（正常人）",
    "Normal": "非AML患者（正常人）",
    "normal": "非AML患者（正常人）",
    "正常": "非AML患者（正常人）",
    "正常人": "非AML患者（正常人）",
    "健康对照": "非AML患者（正常人）",
    "非AML": "非AML患者（正常人）",
    "非AML患者": "非AML患者（正常人）",
}

# 中心颜色映射
CENTER_COLOR_MAP = {
    "PKUPH": "#1f77b4",
    "TAB": "#ff7f0e",
    "BEPH": "#2ca02c",
    "FXH": "#d62728",
    "BJH": "#9467bd",
    "TJMU": "#8c564b",
    "MAIN": "#7f7f7f",
    "Unknown": "#7f7f7f",
}

# 患者大类型点形状映射
PATIENT_TYPE_MARKER_MAP = {
    "AML": "o",
    "非AML": "^",
    "Unknown": "s",
}


# =========================================================
# 4. cell_dict 和细胞中文名称
# =========================================================

cell_dict = {
    "N0": 1, "N": 2, "N1": 3, "N2": 4, "N3": 5, "N4": 6, "N5": 7,
    "E": 8, "B": 9, "M0": 10, "M": 11, "M1": 12, "M2": 13,
    "R": 14, "R1": 15, "R2": 16, "R3": 17,
    "J": 18, "J1": 19, "J2": 20, "J3": 21, "J4": 22,
    "L": 23, "L1": 24, "L2": 25, "L3": 26, "L4": 27,
    "P": 28, "P1": 29, "P2": 30, "P3": 31,
    "B1": 32, "E1": 33, "A": 34, "F": 35, "V": 36, "0": 36
}

# 第二个 sheet 中的展示顺序：
# 每个系统内按照 原始/幼稚细胞在前、成熟细胞在后 排列。
# 注意：label="0" 后续会统一合并到 V，不单独作为一列。
OFFICIAL_CELL_LABEL_ORDER = [
    # 粒细胞系统
    "N0", "N", "N1", "N2", "N3", "N4", "N5", "E", "B",

    # 单核细胞系统
    "M0", "M", "M1", "M2",

    # 红细胞系统
    "R", "R1", "R2", "R3",

    # 巨核细胞系统
    "J", "J1", "J2", "J3", "J4",

    # 淋巴细胞系统
    "L", "L1", "L2", "L3", "L4",

    # 浆细胞系统
    "P", "P1", "P2", "P3",

    # 其他
    "B1", "E1", "A", "F", "V",
]

CELL_LABEL_CN = {
    "N0": "正常原始粒细胞",
    "N": "原始粒细胞(Ⅰ/Ⅱ)",
    "N1": "早幼粒细胞",
    "N2": "中性中幼粒细胞",
    "N3": "中性晚幼粒细胞",
    "N4": "中性杆状核粒细胞",
    "N5": "中性分叶核粒细胞",
    "E": "嗜酸性粒细胞",
    "B": "嗜碱性粒细胞",

    "M0": "正常原始单核细胞",
    "M": "原始单核细胞",
    "M1": "幼稚单核细胞",
    "M2": "成熟单核细胞",

    "R": "原始红细胞",
    "R1": "早幼红细胞",
    "R2": "中幼红细胞",
    "R3": "晚幼红细胞",

    "J": "原始巨核细胞",
    "J1": "幼稚巨核细胞",
    "J2": "颗粒巨核细胞",
    "J3": "产血小板巨核细胞",
    "J4": "裸核细胞",

    "L": "原始淋巴细胞",
    "L1": "幼稚淋巴细胞",
    "L2": "成熟淋巴细胞",
    "L3": "B祖淋巴细胞",
    "L4": "异形淋巴细胞",

    "P": "原始浆细胞",
    "P1": "幼稚浆细胞",
    "P2": "浆细胞",
    "P3": "异常浆细胞",

    "B1": "组织嗜碱细胞",
    "E1": "组织嗜酸细胞",
    "A": "吞噬细胞",
    "F": "脂肪细胞",
    "V": "分类不明细胞",
}


# =========================================================
# 5. 基础工具函数
# =========================================================

def clean_excel_text(x):
    """
    清理 Excel 读入后的文本。
    避免 NaN 显示为 nan，也避免纯数字样本号变成 123.0。
    """
    if pd.isna(x):
        return ""

    if isinstance(x, float):
        if x.is_integer():
            return str(int(x))
        return str(x).strip()

    return str(x).strip()


def normalize_cell_label(label):
    """
    统一细胞标签。
    如果 json 中 label="0"，合并到 V。
    """
    label = str(label).strip()

    if label == "0":
        return "V"

    return label


def infer_dataset_name(directory):
    """
    根据目录名推断数据集名称。
    """
    base = os.path.basename(directory)

    if base.startswith("MAIN"):
        return "MAIN"
    if base.startswith("TJMU"):
        return "TJMU"
    if base.startswith("BJH"):
        return "BJH"
    if base.startswith("FXH"):
        return "FXH"

    return base


def get_patient_id_from_root(root, dataset_dir):
    """
    从当前 root 路径中提取患者 ID。
    """
    rel = os.path.relpath(root, dataset_dir)

    if rel == ".":
        return None

    if ".ipynb_checkpoints" in rel:
        return None

    parts = Path(rel).parts

    if len(parts) == 0:
        return None

    if PATIENT_ID_MODE == "first_level":
        return parts[0]

    if PATIENT_ID_MODE == "full_relative":
        return rel.replace(os.sep, "/")

    raise ValueError(f"Unsupported PATIENT_ID_MODE: {PATIENT_ID_MODE}")


def safe_read_json(json_path):
    """
    安全读取 json。
    """
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            return json.load(f), None
    except Exception as e:
        return None, str(e)


def find_column(df, candidates, required=True):
    """
    在患者信息表中自动寻找列名。
    """
    for c in candidates:
        if c in df.columns:
            return c

    if required:
        raise ValueError(
            f"Cannot find required column. Candidates: {candidates}. "
            f"Available columns: {list(df.columns)}"
        )

    return None


def normalize_center_key(center):
    """
    将中心名称归一化成短 key，便于配色和排序。
    """
    center = str(center).strip()

    if "PKUPH" in center or "北大" in center or "人民" in center:
        return "PKUPH"
    if "TAB" in center or "荻硕" in center:
        return "TAB"
    if "BEPH" in center or "电力" in center:
        return "BEPH"
    if "FXH" in center or "复兴" in center:
        return "FXH"
    if "BJH" in center or "北京医院" in center:
        return "BJH"
    if "TJMU" in center or "同济" in center or "华中" in center:
        return "TJMU"
    if "MAIN" in center:
        return "MAIN"

    return center if center else "Unknown"


def display_center_name(center):
    """
    将中心名称转换成更适合图中和表格中显示的名称。
    """
    center = str(center).strip()

    if center in CENTER_DISPLAY_MAP:
        return CENTER_DISPLAY_MAP[center]

    for key, value in CENTER_DISPLAY_MAP.items():
        if key in center:
            return value

    return center if center else "Unknown"


def normalize_patient_main_type(patient_type):
    """
    表格中统一为 AML / 非AML。
    """
    patient_type = str(patient_type).strip()
    upper_type = patient_type.upper()

    if patient_type in ["非AML", "非AML患者", "非AML患者（正常人）"]:
        return "非AML"

    if "非AML" in patient_type:
        return "非AML"

    if upper_type in ["HC", "NORMAL", "CONTROL", "HEALTHY", "NON-AML", "NONAML"]:
        return "非AML"

    if "正常" in patient_type or "健康" in patient_type or "对照" in patient_type:
        return "非AML"

    if upper_type == "AML":
        return "AML"

    return patient_type if patient_type else "Unknown"


def display_patient_type(patient_type):
    """
    图例中显示：
    AML -> AML患者
    HC / 非AML -> 非AML患者（正常人）
    """
    patient_type = str(patient_type).strip()

    if patient_type in PATIENT_TYPE_DISPLAY_MAP:
        return PATIENT_TYPE_DISPLAY_MAP[patient_type]

    upper_type = patient_type.upper()
    if upper_type in PATIENT_TYPE_DISPLAY_MAP:
        return PATIENT_TYPE_DISPLAY_MAP[upper_type]

    normalized = normalize_patient_main_type(patient_type)

    if normalized == "AML":
        return "AML患者"

    if normalized == "非AML":
        return "非AML患者（正常人）"

    return patient_type if patient_type else "Unknown"


def patient_type_sort_key(patient_type):
    """
    AML 放前，非AML 放后。
    """
    normalized = normalize_patient_main_type(patient_type)

    if normalized == "AML":
        return 1
    if normalized == "非AML":
        return 2

    return 9


def center_sort_key(center):
    """
    控制中心排序。
    """
    center_key = normalize_center_key(center)

    priority = {
        "PKUPH": 1,
        "TAB": 2,
        "BEPH": 3,
        "FXH": 4,
        "BJH": 5,
        "TJMU": 6,
        "MAIN": 7,
        "Unknown": 99,
    }

    return priority.get(center_key, 99)


def patient_id_sort_key(patient_id):
    """
    患者名称排序：
    优先按最后一个数字排序。
    例如：
    BJH-2  < BJH-11
    PKUPH-001 < PKUPH-010
    """
    patient_id = str(patient_id).strip()

    m = re.search(r"(\d+)(?!.*\d)", patient_id)
    if m:
        number = int(m.group(1))
    else:
        number = 10**9

    prefix = re.sub(r"[-_ ]?\d+(?!.*\d)", "", patient_id)

    return prefix, number, patient_id


# =========================================================
# 6. 统计每个患者的 N / M 细胞比例 + 所有类别数量比例
# =========================================================

def collect_patient_label_counts(directories):
    """
    返回患者级统计表。
    每个患者一行，包含：
    1. N_count / N_ratio_pct
    2. M_count / M_ratio_pct
    3. 所有细胞类别的 count / pct
    4. Total_cells_all
    """
    patient_records = {}

    for directory in directories:
        dataset_name = infer_dataset_name(directory)

        print("\n" + "=" * 80)
        print(f"Processing dataset: {dataset_name}")
        print(f"Directory: {directory}")
        print("=" * 80)

        for root, dirs, files in os.walk(directory):
            dirs[:] = [
                d for d in dirs
                if not d.startswith(".")
                and d != "__MACOSX"
                and d != ".ipynb_checkpoints"
            ]

            json_files = [f for f in files if f.lower().endswith(".json")]
            if not json_files:
                continue

            patient_id = get_patient_id_from_root(root, directory)
            if patient_id is None:
                continue

            key = (dataset_name, patient_id)

            if key not in patient_records:
                patient_records[key] = {
                    "Dataset": dataset_name,
                    "Patient_ID": patient_id,
                    "Json_files": 0,
                    "Label_counts": collections.defaultdict(int),
                    "Json_errors": [],
                }

            for filename in json_files:
                json_path = os.path.join(root, filename)
                data, err = safe_read_json(json_path)

                if err is not None:
                    patient_records[key]["Json_errors"].append(f"{json_path}: {err}")
                    continue

                patient_records[key]["Json_files"] += 1

                shapes = data.get("shapes", [])
                for shape in shapes:
                    if POLYGON_ONLY and shape.get("shape_type") != "polygon":
                        continue

                    label = str(shape.get("label", "")).strip()
                    if not label:
                        continue

                    label = normalize_cell_label(label)
                    patient_records[key]["Label_counts"][label] += 1

    rows = []

    for _, info in patient_records.items():
        label_counts = dict(info["Label_counts"])

        defined_total = sum(
            label_counts.get(label, 0)
            for label in OFFICIAL_CELL_LABEL_ORDER
        )

        undefined_labels = {
            label: count
            for label, count in label_counts.items()
            if label not in OFFICIAL_CELL_LABEL_ORDER
        }

        undefined_count = sum(undefined_labels.values())

        total_cells_all = defined_total + undefined_count
        denom = total_cells_all if total_cells_all > 0 else 1

        n_count = sum(label_counts.get(label, 0) for label in N_LABELS)
        m_count = sum(label_counts.get(label, 0) for label in M_LABELS)

        row = {
            "Dataset": info["Dataset"],
            "Patient_ID": info["Patient_ID"],
            "Json_files": info["Json_files"],
            "Total_cells_all": total_cells_all,
            "Defined_cells": defined_total,
            "Undefined_cells": undefined_count,
            "Undefined_labels": ";".join([f"{k}:{v}" for k, v in undefined_labels.items()]),
            "Json_error_count": len(info["Json_errors"]),
            "Json_errors": " | ".join(info["Json_errors"]),
            "N_count": n_count,
            "N_ratio_pct": n_count / denom * 100,
            "M_count": m_count,
            "M_ratio_pct": m_count / denom * 100,
        }

        # 所有类别数量和比例
        for label in OFFICIAL_CELL_LABEL_ORDER:
            count = label_counts.get(label, 0)
            row[f"{label}_count"] = count
            row[f"{label}_pct"] = count / denom * 100

        rows.append(row)

    patient_stats = pd.DataFrame(rows)

    if patient_stats.empty:
        raise RuntimeError("No valid patient-level JSON statistics were found.")

    patient_stats = patient_stats.sort_values(["Dataset", "Patient_ID"]).reset_index(drop=True)

    return patient_stats


# =========================================================
# 7. 读取 patient data 并合并
# =========================================================

def load_patient_info(xlsx_path, sheet_name):
    """
    读取 patient_data_260416.xlsx。
    同时读取：
    1. 患者编号
    2. 样本编号
    3. 患者大类型
    4. 患者小类型
    5. 中心/样本来源
    """
    patient_df = pd.read_excel(xlsx_path, sheet_name=sheet_name)

    id_col = find_column(
        patient_df,
        ["正式编号", "Patient_ID", "patient_id", "编号", "患者编号"]
    )

    sample_id_col = find_column(
        patient_df,
        [
            "样本编号",
            "Sample_ID",
            "sample_id",
            "Sample ID",
            "sample ID",
            "样本号",
            "标本编号",
            "标本号",
            "送检编号",
            "检测编号",
        ],
        required=False
    )

    main_type_col = find_column(
        patient_df,
        ["患者大类型", "Diagnosis", "type", "大类型", "患者类型"]
    )

    subtype_col = find_column(
        patient_df,
        [
            "患者小类型",
            "Subtype",
            "subtype",
            "小类型",
            "具体类型",
            "诊断小类型",
            "患者亚型",
            "亚型",
        ],
        required=False
    )

    source_col = find_column(
        patient_df,
        ["样本来源", "Source", "center", "Center", "中心", "医院"],
        required=False
    )

    keep_cols = [id_col, main_type_col]

    if sample_id_col is not None:
        keep_cols.append(sample_id_col)

    if subtype_col is not None:
        keep_cols.append(subtype_col)

    if source_col is not None:
        keep_cols.append(source_col)

    patient_df = patient_df[keep_cols].copy()

    rename_map = {
        id_col: "Patient_ID",
        main_type_col: "Patient_main_type",
    }

    if sample_id_col is not None:
        rename_map[sample_id_col] = "Sample_ID"
    else:
        patient_df["Sample_ID"] = ""

    if subtype_col is not None:
        rename_map[subtype_col] = "Patient_subtype"
    else:
        patient_df["Patient_subtype"] = ""

    if source_col is not None:
        rename_map[source_col] = "Center"
    else:
        patient_df["Center"] = "Unknown"

    patient_df = patient_df.rename(columns=rename_map)

    patient_df["Patient_ID"] = patient_df["Patient_ID"].apply(clean_excel_text)
    patient_df["Sample_ID"] = patient_df["Sample_ID"].apply(clean_excel_text)
    patient_df["Patient_main_type"] = patient_df["Patient_main_type"].apply(clean_excel_text)
    patient_df["Patient_subtype"] = patient_df["Patient_subtype"].apply(clean_excel_text)
    patient_df["Center"] = patient_df["Center"].apply(clean_excel_text)

    patient_df = patient_df[
        patient_df["Patient_ID"].notna()
        & (patient_df["Patient_ID"] != "")
        & (patient_df["Patient_ID"].str.lower() != "nan")
        & (patient_df["Patient_ID"] != "未使用")
    ].copy()

    # 如果同一患者重复出现，仅保留第一条
    patient_df = patient_df.drop_duplicates(subset=["Patient_ID"], keep="first")

    return patient_df


def merge_patient_info(patient_stats, patient_info):
    """
    将患者级细胞统计与患者信息表合并。
    """
    merged = patient_stats.merge(
        patient_info,
        on="Patient_ID",
        how="left",
        indicator=True
    )

    unmatched = merged[merged["_merge"] != "both"].copy()
    merged = merged.drop(columns=["_merge"])

    merged["Center"] = merged["Center"].fillna(merged["Dataset"])
    merged["Patient_main_type"] = merged["Patient_main_type"].fillna("Unknown")

    if "Sample_ID" not in merged.columns:
        merged["Sample_ID"] = ""

    if "Patient_subtype" not in merged.columns:
        merged["Patient_subtype"] = ""

    merged["Sample_ID"] = merged["Sample_ID"].fillna("").astype(str).str.strip()
    merged["Patient_subtype"] = merged["Patient_subtype"].fillna("").astype(str).str.strip()

    return merged, unmatched


# =========================================================
# 8. 生成 Excel 表格
# =========================================================

def build_patient_nm_ratio_table(merged):
    """
    生成第一个 sheet：
    患者名称、样本编号、中心、大类别、小类别、
    N比例、M比例、N细胞数、M细胞数、细胞总数。

    排序：
    1. 按中心排序
    2. 同一中心内按患者名称末尾编号排序
    """
    table_df = merged.copy()

    table_df["中心排序"] = table_df["Center"].apply(center_sort_key)
    table_df["患者排序"] = table_df["Patient_ID"].apply(patient_id_sort_key)

    table_df["患者名称"] = table_df["Patient_ID"]
    table_df["样本编号"] = table_df["Sample_ID"].fillna("").astype(str).str.strip()
    table_df["中心"] = table_df["Center"].apply(display_center_name)
    table_df["大类别（AML/非AML）"] = table_df["Patient_main_type"].apply(normalize_patient_main_type)
    table_df["小类别"] = table_df["Patient_subtype"].fillna("").astype(str).str.strip()
    table_df["N的比例（%）"] = table_df["N_ratio_pct"].round(3)
    table_df["M的比例（%）"] = table_df["M_ratio_pct"].round(3)
    table_df["N细胞数"] = table_df["N_count"].astype(int)
    table_df["M细胞数"] = table_df["M_count"].astype(int)
    table_df["细胞总数"] = table_df["Total_cells_all"].astype(int)

    table_df = table_df.sort_values(
        by=["中心排序", "患者排序", "患者名称"],
        ascending=[True, True, True]
    )

    final_table = table_df[
        [
            "患者名称",
            "样本编号",
            "中心",
            "大类别（AML/非AML）",
            "小类别",
            "N的比例（%）",
            "M的比例（%）",
            "N细胞数",
            "M细胞数",
            "细胞总数",
        ]
    ].reset_index(drop=True)

    return final_table


def build_all_cell_category_table(merged):
    """
    生成第二个 sheet：
    每个患者所有细胞类别的数量和比例。

    列顺序：
    1. 患者基本信息
    2. 粒系：原始/幼稚 -> 成熟
    3. 单核系：原始/幼稚 -> 成熟
    4. 红系：原始 -> 晚幼
    5. 巨核系：原始 -> 成熟
    6. 淋巴系：原始 -> 成熟/异常
    7. 浆细胞系：原始 -> 成熟/异常
    8. 其他类别
    """
    table_df = merged.copy()

    table_df["中心排序"] = table_df["Center"].apply(center_sort_key)
    table_df["患者排序"] = table_df["Patient_ID"].apply(patient_id_sort_key)

    table_df["患者名称"] = table_df["Patient_ID"]
    table_df["样本编号"] = table_df["Sample_ID"].fillna("").astype(str).str.strip()
    table_df["中心"] = table_df["Center"].apply(display_center_name)
    table_df["大类别（AML/非AML）"] = table_df["Patient_main_type"].apply(normalize_patient_main_type)
    table_df["小类别"] = table_df["Patient_subtype"].fillna("").astype(str).str.strip()
    table_df["细胞总数"] = table_df["Total_cells_all"].astype(int)

    output_cols = [
        "患者名称",
        "样本编号",
        "中心",
        "大类别（AML/非AML）",
        "小类别",
        "细胞总数",
    ]

    for label in OFFICIAL_CELL_LABEL_ORDER:
        cn_name = CELL_LABEL_CN.get(label, label)

        count_col = f"{cn_name}（{label}）数量"
        pct_col = f"{cn_name}（{label}）比例（%）"

        table_df[count_col] = table_df[f"{label}_count"].fillna(0).astype(int)
        table_df[pct_col] = table_df[f"{label}_pct"].fillna(0).round(3)

        output_cols.extend([count_col, pct_col])

    table_df = table_df.sort_values(
        by=["中心排序", "患者排序", "患者名称"],
        ascending=[True, True, True]
    )

    final_table = table_df[output_cols].reset_index(drop=True)

    return final_table


def style_excel_sheet(ws, table_name):
    """
    给 Excel sheet 设置样式。
    """
    ws.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="000000")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    # 自动设置列宽
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        header_value = str(ws.cell(row=1, column=col_idx).value)

        if col_idx == 1:
            width = 18
        elif "样本编号" in header_value:
            width = 22
        elif "中心" in header_value:
            width = 28
        elif "类别" in header_value:
            width = 18
        elif "比例" in header_value:
            width = 16
        elif "数量" in header_value or "细胞数" in header_value or "细胞总数" in header_value:
            width = 14
        else:
            width = 16

        ws.column_dimensions[col_letter].width = width

    # 所有比例列保留 3 位小数
    for col_idx in range(1, ws.max_column + 1):
        header_value = str(ws.cell(row=1, column=col_idx).value)
        if "比例" in header_value:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0.000"

    # 添加 Excel 表格样式
    max_row = ws.max_row
    max_col = ws.max_column
    table_ref = f"A1:{get_column_letter(max_col)}{max_row}"

    excel_table = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    excel_table.tableStyleInfo = style
    ws.add_table(excel_table)


def save_patient_table_excel(patient_nm_table, all_category_table, out_path):
    """
    保存 Excel：
    Sheet1：患者NM比例表
    Sheet2：所有类别数量比例
    """
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        patient_nm_table.to_excel(
            writer,
            index=False,
            sheet_name="患者NM比例表"
        )

        all_category_table.to_excel(
            writer,
            index=False,
            sheet_name="所有类别数量比例"
        )

    wb = load_workbook(out_path)

    style_excel_sheet(
        wb["患者NM比例表"],
        table_name="PatientNMRatioTable"
    )

    style_excel_sheet(
        wb["所有类别数量比例"],
        table_name="AllCellCategoryTable"
    )

    wb.save(out_path)


# =========================================================
# 9. 作图：二维 N-M 比例散点图
# =========================================================

def make_nm_ratio_2d_scatter_plot(
    df,
    title,
    out_prefix,
):
    """
    二维散点图：
    横轴：N 细胞比例
    纵轴：M 细胞比例

    点颜色：中心
    点形状：患者大类型，AML / 非AML
    """
    plot_df = df.copy()
    plot_df = plot_df[
        plot_df["N_ratio_pct"].notna()
        & plot_df["M_ratio_pct"].notna()
    ].copy()

    if len(plot_df) == 0:
        print(f"⚠️ No valid data for plot: {title}")
        return None, None, None

    plot_df["Center_key"] = plot_df["Center"].apply(normalize_center_key)
    plot_df["Center_display"] = plot_df["Center"].apply(display_center_name)
    plot_df["Patient_type_norm"] = plot_df["Patient_main_type"].apply(normalize_patient_main_type)
    plot_df["Patient_type_display"] = plot_df["Patient_main_type"].apply(display_patient_type)
    plot_df["Center_sort"] = plot_df["Center"].apply(center_sort_key)
    plot_df["Patient_type_sort"] = plot_df["Patient_main_type"].apply(patient_type_sort_key)

    fig, ax = plt.subplots(figsize=(FIG_WIDTH, FIG_HEIGHT), dpi=FIG_DPI)

    group_cols = [
        "Center_key",
        "Center_display",
        "Patient_type_norm",
        "Patient_type_display",
    ]

    group_df = (
        plot_df[group_cols + ["Center_sort", "Patient_type_sort"]]
        .drop_duplicates()
        .sort_values(["Center_sort", "Patient_type_sort", "Center_key", "Patient_type_norm"])
    )

    for _, g in group_df.iterrows():
        center_key = g["Center_key"]
        center_display = g["Center_display"]
        patient_type_norm = g["Patient_type_norm"]
        patient_type_display = g["Patient_type_display"]

        sub = plot_df[
            (plot_df["Center_key"] == center_key)
            & (plot_df["Patient_type_norm"] == patient_type_norm)
        ].copy()

        if len(sub) == 0:
            continue

        color = CENTER_COLOR_MAP.get(center_key, "#7f7f7f")
        marker = PATIENT_TYPE_MARKER_MAP.get(patient_type_norm, "s")

        label = f"{center_display} | {patient_type_display}"

        ax.scatter(
            sub["N_ratio_pct"],
            sub["M_ratio_pct"],
            s=POINT_SIZE,
            alpha=POINT_ALPHA,
            c=color,
            marker=marker,
            edgecolors="black",
            linewidths=0.45,
            label=label,
        )

        if ANNOTATE_PATIENT_ID:
            for _, row in sub.iterrows():
                ax.text(
                    row["N_ratio_pct"] + 0.2,
                    row["M_ratio_pct"] + 0.2,
                    str(row["Patient_ID"]),
                    fontsize=ANNOTATE_FONTSIZE,
                    alpha=0.75,
                )

    ax.set_title(title, fontsize=15, pad=16)
    ax.set_xlabel("N 细胞比例（%）", fontsize=12)
    ax.set_ylabel("M 细胞比例（%）", fontsize=12)
    ax.grid(True, linestyle="--", alpha=0.35)

    # 5% 参考线
    if SHOW_REF_LINE:
        ax.axvline(
            x=REF_RATIO,
            linestyle="--",
            linewidth=1.4,
            alpha=0.85,
            color="red"
        )

        ax.axhline(
            y=REF_RATIO,
            linestyle="--",
            linewidth=1.4,
            alpha=0.85,
            color="red"
        )

    # 保证 5% 参考线和全部点都能显示
    # 同时扩展 0 点附近坐标轴，避免 x=0 或 y=0 的点被裁切
    max_x = max(plot_df["N_ratio_pct"].max(), REF_RATIO)
    max_y = max(plot_df["M_ratio_pct"].max(), REF_RATIO)

    x_upper = max_x * 1.15 if max_x > 0 else 10
    y_upper = max_y * 1.15 if max_y > 0 else 10

    x_pad = max(x_upper * 0.03, 0.8)
    y_pad = max(y_upper * 0.03, 0.8)

    ax.set_xlim(-x_pad, x_upper)
    ax.set_ylim(-y_pad, y_upper)

    # 5% 标签在坐标轴范围确定后再添加
    if SHOW_REF_LINE:
        ax.text(
            REF_RATIO,
            y_upper * 0.98,
            "N=5%",
            va="top",
            ha="left",
            fontsize=10,
            color="red"
        )

        ax.text(
            x_upper * 0.98,
            REF_RATIO,
            "M=5%",
            va="bottom",
            ha="right",
            fontsize=10,
            color="red"
        )

    ax.legend(
        title="中心 | 患者类型",
        fontsize=8,
        title_fontsize=9,
        loc="center left",
        bbox_to_anchor=(1.02, 0.5),
        frameon=False,
        markerscale=0.9,
    )

    fig.subplots_adjust(
        left=0.10,
        right=0.70,
        top=0.90,
        bottom=0.12
    )

    png_path = f"{out_prefix}.png"
    pdf_path = f"{out_prefix}.pdf"
    svg_path = f"{out_prefix}.svg"

    fig.savefig(png_path, dpi=FIG_DPI, bbox_inches="tight")
    fig.savefig(pdf_path, bbox_inches="tight")
    fig.savefig(svg_path, bbox_inches="tight")
    plt.close(fig)

    return png_path, pdf_path, svg_path


# =========================================================
# 10. 主程序
# =========================================================

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 中文字体已由你的环境统一配置，这里不再设置字体
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["pdf.fonttype"] = 42
    plt.rcParams["ps.fonttype"] = 42
    plt.rcParams["svg.fonttype"] = "path"

    print("=" * 80)
    print("Step 1: 统计每个患者的 N / M 细胞比例和所有类别比例")
    print("=" * 80)

    patient_stats = collect_patient_label_counts(DIRECTORIES)
    print(f"共统计患者数：{len(patient_stats)}")

    print("=" * 80)
    print("Step 2: 读取 patient data 并合并样本编号、中心、大类别和小类别")
    print("=" * 80)

    patient_info = load_patient_info(PATIENT_INFO_XLSX, PATIENT_INFO_SHEET)
    merged, unmatched = merge_patient_info(patient_stats, patient_info)

    print(f"患者信息表记录数：{len(patient_info)}")
    print(f"未匹配患者数：{len(unmatched)}")

    if len(unmatched) > 0:
        print("⚠️ 有患者未在 patient data 表中匹配到信息。")
        print("这些患者的样本编号/小类别可能为空，中心会用 Dataset 代替。")

    print("=" * 80)
    print("Step 3: 输出 Excel 表格")
    print("=" * 80)

    patient_table = build_patient_nm_ratio_table(merged)
    all_category_table = build_all_cell_category_table(merged)

    table_xlsx_path = os.path.join(
        OUTPUT_DIR,
        "patient_NM_ratio_clear_table.xlsx"
    )

    save_patient_table_excel(
        patient_nm_table=patient_table,
        all_category_table=all_category_table,
        out_path=table_xlsx_path
    )

    print(f"患者 N/M 比例表 + 所有类别数量比例表：{table_xlsx_path}")

    print("=" * 80)
    print("Step 4: 输出二维 N-M 比例散点图")
    print("=" * 80)

    nm_plot_prefix = os.path.join(
        OUTPUT_DIR,
        "center_patient_main_type_NM_ratio_2d_scatter"
    )

    nm_plot_paths = make_nm_ratio_2d_scatter_plot(
        merged,
        title="各中心不同患者大类型的 N 与 M 细胞比例分布",
        out_prefix=nm_plot_prefix
    )

    print("二维 N-M 比例散点图：")
    print(nm_plot_paths)

    print("=" * 80)
    print("完成！仅输出以下内容：")
    print(f"1. {table_xlsx_path}")
    print(f"2. {nm_plot_paths}")
    print("=" * 80)


if __name__ == "__main__":
    main()