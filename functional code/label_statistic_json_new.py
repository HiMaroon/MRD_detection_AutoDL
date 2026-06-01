import os
import re
import json
import collections
from pathlib import Path
import textwrap

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


# =========================================================
# 1. 配置区
# =========================================================

DIRECTORIES = [
    "/root/autodl-tmp/data/MAIN_imgs_260323",
    "/root/autodl-tmp/data/TJMU_imgs_260416",
    "/root/autodl-tmp/data/BJH_imgs_260211",
    "/root/autodl-tmp/data/FXH_imgs_noALL_260318",
]

PATIENT_INFO_XLSX = "/root/autodl-tmp/data/patient_data_260416.xlsx"
PATIENT_INFO_SHEET = "总表"

OUTPUT_DIR = "/root/autodl-tmp/data/patient_N_ratio_distribution_260416"

# 是否只统计 polygon 标注
POLYGON_ONLY = True

# 患者 ID 提取方式：
# first_level: 每个主目录下第一层子文件夹为患者，例如 MAIN_imgs_260323/PKUPH-001/xxx.json
# full_relative: 使用完整相对路径作为患者 ID
PATIENT_ID_MODE = "first_level"

# N 细胞比例定义
# 默认只统计精确标签 "N"
N_LABELS = ["N"]


# =========================================================
# 2. 作图参数
# =========================================================

FIG_DPI = 350
FIG_HEIGHT = 7.2
FIG_WIDTH_PER_GROUP = 1.9
MIN_FIG_WIDTH = 16

# X 轴每个分组之间的间距
X_SPACING = 1.8

# X 轴标签旋转角度
X_TICK_ROTATION = 35

CENTER_WRAP_WIDTH = 10
TYPE_WRAP_WIDTH = 14

# 散点抖动范围
JITTER_SD = 0.10

# 是否在每个分组上方标注 n
SHOW_N_LABEL = True

# 5% 参考线
SHOW_REF_LINE = True
REF_RATIO = 5.0


# =========================================================
# 3. 显示名称映射
# =========================================================

CENTER_DISPLAY_MAP = {
    "PKUPH": "北大人民医院\nPKUPH",
    "北大人民医院": "北大人民医院\nPKUPH",
    "北京大学人民医院": "北大人民医院\nPKUPH",

    "TAB": "荻硕贝肯\nTAB",
    "荻硕贝肯": "荻硕贝肯\nTAB",

    "BEPH": "北京电力医院\nBEPH",
    "北京电力医院": "北京电力医院\nBEPH",

    "FXH": "复兴医院\nFXH",
    "首都医科大学附属复兴医院": "复兴医院\nFXH",

    "BJH": "北京医院\nBJH",
    "北京医院血液实验室": "北京医院血液实验室\nBJH",

    "TJMU": "华中科技大学同济医学院\nTJMU",
    "华中科技大学同济医学院": "华中科技大学同济医学院\nTJMU",

    "MAIN": "MAIN",
}

PATIENT_TYPE_DISPLAY_MAP = {
    "AML": "AML患者",
    "HC": "非AML患者（正常人）",
    "NORMAL": "非AML患者（正常人）",
    "Normal": "非AML患者（正常人）",
    "normal": "非AML患者（正常人）",
    "正常": "非AML患者（正常人）",
    "正常人": "非AML患者（正常人）",
    "健康对照": "非AML患者（正常人）",
    "非AML": "非AML患者（正常人）",
    "非AML患者": "非AML患者（正常人）",
}


# =========================================================
# 4. cell_dict
# =========================================================

cell_dict = {
    "N0": 1, "N": 2, "N1": 3, "N2": 4, "N3": 5, "N4": 6, "N5": 7,
    "E": 8, "B": 9, "M0": 10, "M": 11, "M1": 12, "M2": 13,
    "R": 14, "R1": 15, "R2": 16, "R3": 17,
    "J": 18, "J1": 19, "J2": 20, "J3": 21, "J4": 22,
    "L": 23, "L1": 24, "L2": 25, "L3": 26, "L4": 27,
    "P": 28, "P1": 29, "P2": 30, "P3": 31,
    "B1": 32, "E1": 33, "A": 34, "F": 35, "V": 36, "0": 36
}

SORTED_LABELS = sorted(cell_dict.keys(), key=lambda x: cell_dict[x])


# =========================================================
# 5. 基础工具函数
# =========================================================

def clean_excel_text(x):
    """
    清理 Excel 读入后的文本。
    避免 NaN 显示为 nan，也避免纯数字样本号变成 123.0。
    """
    if pd.isna(x):
        return ""

    if isinstance(x, float):
        if x.is_integer():
            return str(int(x))
        return str(x).strip()

    return str(x).strip()


def infer_dataset_name(directory):
    """
    根据目录名推断数据集名称。
    """
    base = os.path.basename(directory)

    if base.startswith("MAIN"):
        return "MAIN"
    if base.startswith("TJMU"):
        return "TJMU"
    if base.startswith("BJH"):
        return "BJH"
    if base.startswith("FXH"):
        return "FXH"

    return base


def get_patient_id_from_root(root, dataset_dir):
    """
    从当前 root 路径中提取患者 ID。
    """
    rel = os.path.relpath(root, dataset_dir)

    if rel == ".":
        return None

    if ".ipynb_checkpoints" in rel:
        return None

    parts = Path(rel).parts

    if len(parts) == 0:
        return None

    if PATIENT_ID_MODE == "first_level":
        return parts[0]

    if PATIENT_ID_MODE == "full_relative":
        return rel.replace(os.sep, "/")

    raise ValueError(f"Unsupported PATIENT_ID_MODE: {PATIENT_ID_MODE}")


def safe_read_json(json_path):
    """
    安全读取 json。
    """
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            return json.load(f), None
    except Exception as e:
        return None, str(e)


def find_column(df, candidates, required=True):
    """
    在患者信息表中自动寻找列名。
    """
    for c in candidates:
        if c in df.columns:
            return c

    if required:
        raise ValueError(
            f"Cannot find required column. Candidates: {candidates}. "
            f"Available columns: {list(df.columns)}"
        )

    return None


def wrap_text(s, width=10):
    """
    用于 X 轴标签换行。
    """
    s = str(s)

    if "\n" in s:
        return s

    return "\n".join(textwrap.wrap(s, width=width)) if len(s) > width else s


def display_center_name(center):
    """
    将中心名称转换成更适合图中显示的短标签。
    """
    center = str(center).strip()

    if center in CENTER_DISPLAY_MAP:
        return CENTER_DISPLAY_MAP[center]

    for key, value in CENTER_DISPLAY_MAP.items():
        if key in center:
            return value

    return wrap_text(center, CENTER_WRAP_WIDTH)


def display_center_name_for_table(center):
    """
    表格中使用单行中心名称。
    """
    return display_center_name(center).replace("\n", " ")


def normalize_patient_main_type(patient_type):
    """
    表格中统一为 AML / 非AML。
    """
    patient_type = str(patient_type).strip()
    upper_type = patient_type.upper()

    if patient_type in ["非AML", "非AML患者", "非AML患者（正常人）"]:
        return "非AML"

    if "非AML" in patient_type:
        return "非AML"

    if upper_type in ["HC", "NORMAL", "CONTROL", "HEALTHY", "NON-AML", "NONAML"]:
        return "非AML"

    if "正常" in patient_type or "健康" in patient_type or "对照" in patient_type:
        return "非AML"

    if upper_type == "AML":
        return "AML"

    return patient_type


def display_patient_type(patient_type):
    """
    横轴中显示：
    AML -> AML患者
    HC / 非AML -> 非AML患者（正常人）
    """
    patient_type = str(patient_type).strip()

    if patient_type in PATIENT_TYPE_DISPLAY_MAP:
        return PATIENT_TYPE_DISPLAY_MAP[patient_type]

    upper_type = patient_type.upper()
    if upper_type in PATIENT_TYPE_DISPLAY_MAP:
        return PATIENT_TYPE_DISPLAY_MAP[upper_type]

    normalized = normalize_patient_main_type(patient_type)

    if normalized == "AML":
        return "AML患者"

    if normalized == "非AML":
        return "非AML患者（正常人）"

    return patient_type


def patient_type_sort_key(patient_type):
    """
    作图时 AML 放前，非AML 放后。
    """
    normalized = normalize_patient_main_type(patient_type)

    if normalized == "AML":
        return 1
    if normalized == "非AML":
        return 2

    return 9


def center_sort_key(center):
    """
    控制中心排序。
    """
    center = str(center)

    priority_rules = [
        ("PKUPH", 1), ("北大", 1), ("人民", 1),
        ("TAB", 2), ("荻硕", 2),
        ("BEPH", 3), ("电力", 3),
        ("FXH", 4), ("复兴", 4),
        ("BJH", 5), ("北京医院", 5),
        ("TJMU", 6), ("同济", 6), ("华中", 6),
        ("MAIN", 7),
    ]

    for keyword, score in priority_rules:
        if keyword in center:
            return score

    return 99


def patient_id_sort_key(patient_id):
    """
    患者名称排序：
    优先按最后一个数字排序。
    例如：
    BJH-2  < BJH-11
    PKUPH-001 < PKUPH-010
    """
    patient_id = str(patient_id).strip()

    # 提取最后一段数字
    m = re.search(r"(\d+)(?!.*\d)", patient_id)
    if m:
        number = int(m.group(1))
    else:
        number = 10**9

    # 去掉末尾数字后的前缀作为辅助排序
    prefix = re.sub(r"[-_ ]?\d+(?!.*\d)", "", patient_id)

    return prefix, number, patient_id


# =========================================================
# 6. 统计每个患者的 N 细胞比例
# =========================================================

def collect_patient_label_counts(directories):
    """
    返回患者级统计表。
    每个患者一行，包含：
    1. N_count
    2. N_ratio_pct
    3. Total_cells_all
    """
    patient_records = {}

    for directory in directories:
        dataset_name = infer_dataset_name(directory)

        print("\n" + "=" * 80)
        print(f"Processing dataset: {dataset_name}")
        print(f"Directory: {directory}")
        print("=" * 80)

        for root, dirs, files in os.walk(directory):
            dirs[:] = [
                d for d in dirs
                if not d.startswith(".")
                and d != "__MACOSX"
                and d != ".ipynb_checkpoints"
            ]

            json_files = [f for f in files if f.lower().endswith(".json")]
            if not json_files:
                continue

            patient_id = get_patient_id_from_root(root, directory)
            if patient_id is None:
                continue

            key = (dataset_name, patient_id)

            if key not in patient_records:
                patient_records[key] = {
                    "Dataset": dataset_name,
                    "Patient_ID": patient_id,
                    "Json_files": 0,
                    "Label_counts": collections.defaultdict(int),
                    "Json_errors": [],
                }

            for filename in json_files:
                json_path = os.path.join(root, filename)
                data, err = safe_read_json(json_path)

                if err is not None:
                    patient_records[key]["Json_errors"].append(f"{json_path}: {err}")
                    continue

                patient_records[key]["Json_files"] += 1

                shapes = data.get("shapes", [])
                for shape in shapes:
                    if POLYGON_ONLY and shape.get("shape_type") != "polygon":
                        continue

                    label = str(shape.get("label", "")).strip()
                    if not label:
                        continue

                    patient_records[key]["Label_counts"][label] += 1

    rows = []

    for _, info in patient_records.items():
        label_counts = dict(info["Label_counts"])

        defined_total = sum(label_counts.get(label, 0) for label in SORTED_LABELS)

        undefined_labels = {
            label: count
            for label, count in label_counts.items()
            if label not in cell_dict
        }

        undefined_count = sum(undefined_labels.values())

        total_cells_all = defined_total + undefined_count
        denom = total_cells_all if total_cells_all > 0 else 1

        n_count = sum(label_counts.get(label, 0) for label in N_LABELS)
        n_ratio_pct = n_count / denom * 100

        row = {
            "Dataset": info["Dataset"],
            "Patient_ID": info["Patient_ID"],
            "Json_files": info["Json_files"],
            "Total_cells_all": total_cells_all,
            "Defined_cells": defined_total,
            "Undefined_cells": undefined_count,
            "Undefined_labels": ";".join([f"{k}:{v}" for k, v in undefined_labels.items()]),
            "Json_error_count": len(info["Json_errors"]),
            "Json_errors": " | ".join(info["Json_errors"]),
            "N_count": n_count,
            "N_ratio_pct": n_ratio_pct,
        }

        rows.append(row)

    patient_stats = pd.DataFrame(rows)

    if patient_stats.empty:
        raise RuntimeError("No valid patient-level JSON statistics were found.")

    patient_stats = patient_stats.sort_values(["Dataset", "Patient_ID"]).reset_index(drop=True)

    return patient_stats


# =========================================================
# 7. 读取 patient data 并合并
# =========================================================

def load_patient_info(xlsx_path, sheet_name):
    """
    读取 patient_data_260416.xlsx。
    同时读取：
    1. 患者编号
    2. 样本编号
    3. 患者大类型
    4. 中心/样本来源
    """
    patient_df = pd.read_excel(xlsx_path, sheet_name=sheet_name)

    id_col = find_column(
        patient_df,
        ["正式编号", "Patient_ID", "patient_id", "编号", "患者编号"]
    )

    sample_id_col = find_column(
        patient_df,
        [
            "样本编号",
            "Sample_ID",
            "sample_id",
            "Sample ID",
            "sample ID",
            "样本号",
            "标本编号",
            "标本号",
            "送检编号",
            "检测编号",
        ],
        required=False
    )

    main_type_col = find_column(
        patient_df,
        ["患者大类型", "Diagnosis", "type", "大类型", "患者类型"]
    )

    source_col = find_column(
        patient_df,
        ["样本来源", "Source", "center", "Center", "中心", "医院"],
        required=False
    )

    keep_cols = [id_col, main_type_col]

    if sample_id_col is not None:
        keep_cols.append(sample_id_col)

    if source_col is not None:
        keep_cols.append(source_col)

    patient_df = patient_df[keep_cols].copy()

    rename_map = {
        id_col: "Patient_ID",
        main_type_col: "Patient_main_type",
    }

    if sample_id_col is not None:
        rename_map[sample_id_col] = "Sample_ID"
    else:
        patient_df["Sample_ID"] = ""

    if source_col is not None:
        rename_map[source_col] = "Center"
    else:
        patient_df["Center"] = "Unknown"

    patient_df = patient_df.rename(columns=rename_map)

    patient_df["Patient_ID"] = patient_df["Patient_ID"].apply(clean_excel_text)
    patient_df["Sample_ID"] = patient_df["Sample_ID"].apply(clean_excel_text)
    patient_df["Patient_main_type"] = patient_df["Patient_main_type"].apply(clean_excel_text)
    patient_df["Center"] = patient_df["Center"].apply(clean_excel_text)

    # 去除无效编号
    patient_df = patient_df[
        patient_df["Patient_ID"].notna()
        & (patient_df["Patient_ID"] != "")
        & (patient_df["Patient_ID"].str.lower() != "nan")
        & (patient_df["Patient_ID"] != "未使用")
    ].copy()

    # 如果同一患者重复出现，仅保留第一条
    patient_df = patient_df.drop_duplicates(subset=["Patient_ID"], keep="first")

    return patient_df


def merge_patient_info(patient_stats, patient_info):
    """
    将患者级细胞统计与患者信息表合并。
    """
    merged = patient_stats.merge(
        patient_info,
        on="Patient_ID",
        how="left",
        indicator=True
    )

    unmatched = merged[merged["_merge"] != "both"].copy()

    merged = merged.drop(columns=["_merge"])

    merged["Center"] = merged["Center"].fillna(merged["Dataset"])
    merged["Patient_main_type"] = merged["Patient_main_type"].fillna("Unknown")

    if "Sample_ID" not in merged.columns:
        merged["Sample_ID"] = ""

    merged["Sample_ID"] = merged["Sample_ID"].fillna("").astype(str).str.strip()

    return merged, unmatched


# =========================================================
# 8. 生成清晰患者表
# =========================================================

def build_patient_n_ratio_table(merged):
    """
    生成最终表格：
    患者名称、样本编号、中心、大类别（AML/非AML）、N的比例、N细胞数、细胞总数。

    排序：
    1. 按中心排序
    2. 同一中心内按患者名称末尾编号排序
    """
    table_df = merged.copy()

    table_df["中心排序"] = table_df["Center"].apply(center_sort_key)
    table_df["患者排序"] = table_df["Patient_ID"].apply(patient_id_sort_key)

    table_df["患者名称"] = table_df["Patient_ID"]
    table_df["样本编号"] = table_df["Sample_ID"].fillna("").astype(str).str.strip()
    table_df["中心"] = table_df["Center"].apply(display_center_name_for_table)
    table_df["大类别（AML/非AML）"] = table_df["Patient_main_type"].apply(normalize_patient_main_type)
    table_df["N的比例（%）"] = table_df["N_ratio_pct"].round(3)
    table_df["N细胞数"] = table_df["N_count"].astype(int)
    table_df["细胞总数"] = table_df["Total_cells_all"].astype(int)

    table_df = table_df.sort_values(
        by=["中心排序", "患者排序", "患者名称"],
        ascending=[True, True, True]
    )

    final_table = table_df[
        [
            "患者名称",
            "样本编号",
            "中心",
            "大类别（AML/非AML）",
            "N的比例（%）",
            "N细胞数",
            "细胞总数",
        ]
    ].reset_index(drop=True)

    return final_table


def save_patient_table_excel(table_df, out_path):
    """
    保存一个清晰可读的 Excel 表格。
    """
    table_df.to_excel(out_path, index=False, sheet_name="患者N比例表")

    wb = load_workbook(out_path)
    ws = wb["患者N比例表"]

    # 冻结首行
    ws.freeze_panes = "A2"

    # 样式
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="000000")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # 设置列宽
    column_widths = {
        "A": 18,  # 患者名称
        "B": 22,  # 样本编号
        "C": 28,  # 中心
        "D": 20,  # 大类别
        "E": 16,  # N比例
        "F": 14,  # N细胞数
        "G": 14,  # 细胞总数
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 添加 Excel 表格样式
    max_row = ws.max_row
    table_ref = f"A1:G{max_row}"

    excel_table = Table(displayName="PatientNRatioTable", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    excel_table.tableStyleInfo = style
    ws.add_table(excel_table)

    # N比例列数字格式
    for row in range(2, max_row + 1):
        ws[f"E{row}"].number_format = "0.000"

    wb.save(out_path)


# =========================================================
# 9. 作图：只画中心 + 患者大类型的 N 比例散点图
# =========================================================

def make_center_main_type_n_ratio_scatter_plot(
    df,
    title,
    xlabel,
    out_prefix,
):
    """
    只输出 center_patient_main_type_N_ratio_scatter 系列图片。
    每个点代表一个患者。
    """
    plot_df = df.copy()
    plot_df = plot_df[plot_df["N_ratio_pct"].notna()].copy()

    if len(plot_df) == 0:
        print(f"⚠️ No valid data for plot: {title}")
        return None, None, None

    plot_df["Group_label"] = (
        plot_df["Center"].apply(display_center_name)
        + "\n"
        + plot_df["Patient_main_type"].astype(str).apply(
            lambda x: wrap_text(display_patient_type(x), TYPE_WRAP_WIDTH)
        )
    )

    plot_df["Center_sort"] = plot_df["Center"].apply(center_sort_key)
    plot_df["Patient_type_sort"] = plot_df["Patient_main_type"].apply(patient_type_sort_key)

    group_order_df = (
        plot_df[["Center", "Center_sort", "Patient_main_type", "Patient_type_sort", "Group_label"]]
        .drop_duplicates()
        .sort_values(["Center_sort", "Patient_type_sort", "Patient_main_type"])
    )

    labels = group_order_df["Group_label"].tolist()
    positions = np.arange(1, len(labels) + 1) * X_SPACING

    fig_width = max(MIN_FIG_WIDTH, len(labels) * FIG_WIDTH_PER_GROUP)

    fig, ax = plt.subplots(figsize=(fig_width, FIG_HEIGHT), dpi=FIG_DPI)

    rng = np.random.default_rng(42)

    for pos, label in zip(positions, labels):
        sub = plot_df.loc[plot_df["Group_label"] == label].copy()
        vals = sub["N_ratio_pct"].dropna().values

        if len(vals) == 0:
            continue

        jitter = rng.normal(0, JITTER_SD, size=len(vals))

        ax.scatter(
            np.full(len(vals), pos) + jitter,
            vals,
            s=42,
            alpha=0.85,
            edgecolors="none"
        )

        if SHOW_N_LABEL:
            y_max = max(plot_df["N_ratio_pct"].max(), REF_RATIO, 1)
            ax.text(
                pos,
                y_max * 1.05,
                f"n={len(vals)}",
                ha="center",
                va="bottom",
                fontsize=9,
            )

    ax.set_xticks(positions)
    ax.set_xticklabels(
        labels,
        rotation=X_TICK_ROTATION,
        ha="right",
        rotation_mode="anchor",
        fontsize=10
    )

    ax.set_title(title, fontsize=15, pad=18)
    ax.set_xlabel(xlabel, fontsize=12)
    ax.set_ylabel("N 细胞比例（%）", fontsize=12)

    if SHOW_REF_LINE:
        ax.axhline(
            y=REF_RATIO,
            linestyle="--",
            linewidth=1.6,
            alpha=0.85,
            color="red"
        )

        ax.text(
            positions[-1] + X_SPACING * 0.25,
            REF_RATIO + 2,
            "5%",
            va="center",
            ha="left",
            fontsize=10,
            color="red"
        )

    ax.grid(axis="y", linestyle="--", alpha=0.35)

    y_upper = max(plot_df["N_ratio_pct"].max(), REF_RATIO) * 1.15
    ax.set_ylim(0, y_upper)

    ax.margins(x=0.04)

    fig.subplots_adjust(
        left=0.07,
        right=0.98,
        top=0.88,
        bottom=0.35
    )

    png_path = f"{out_prefix}.png"
    pdf_path = f"{out_prefix}.pdf"
    svg_path = f"{out_prefix}.svg"

    fig.savefig(png_path, dpi=FIG_DPI, bbox_inches="tight")
    fig.savefig(pdf_path, bbox_inches="tight")
    fig.savefig(svg_path, bbox_inches="tight")
    plt.close(fig)

    return png_path, pdf_path, svg_path


# =========================================================
# 10. 主程序
# =========================================================

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 中文字体已由你的环境统一配置，这里不再设置字体
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["pdf.fonttype"] = 42
    plt.rcParams["ps.fonttype"] = 42
    plt.rcParams["svg.fonttype"] = "path"

    print("=" * 80)
    print("Step 1: 统计每个患者的 N 细胞比例")
    print("=" * 80)

    patient_stats = collect_patient_label_counts(DIRECTORIES)
    print(f"共统计患者数：{len(patient_stats)}")

    print("=" * 80)
    print("Step 2: 读取 patient data 并合并样本编号、中心和大类别")
    print("=" * 80)

    patient_info = load_patient_info(PATIENT_INFO_XLSX, PATIENT_INFO_SHEET)
    merged, unmatched = merge_patient_info(patient_stats, patient_info)

    print(f"患者信息表记录数：{len(patient_info)}")
    print(f"未匹配患者数：{len(unmatched)}")

    if len(unmatched) > 0:
        print("⚠️ 有患者未在 patient data 表中匹配到信息。")
        print("这些患者的样本编号可能为空，中心会用 Dataset 代替。")

    print("=" * 80)
    print("Step 3: 输出清晰患者 N 比例表")
    print("=" * 80)

    patient_table = build_patient_n_ratio_table(merged)

    table_xlsx_path = os.path.join(
        OUTPUT_DIR,
        "patient_N_ratio_clear_table.xlsx"
    )

    save_patient_table_excel(patient_table, table_xlsx_path)

    print(f"患者 N 比例表格：{table_xlsx_path}")

    print("=" * 80)
    print("Step 4: 输出 center_patient_main_type_N_ratio_scatter 系列图片")
    print("=" * 80)

    n_main_plot_prefix = os.path.join(
        OUTPUT_DIR,
        "center_patient_main_type_N_ratio_scatter"
    )

    n_main_paths = make_center_main_type_n_ratio_scatter_plot(
        merged,
        title="各中心不同患者大类型的 N 细胞比例分布",
        xlabel="中心及患者大类型",
        out_prefix=n_main_plot_prefix
    )

    print("center_patient_main_type_N_ratio_scatter 系列图片：")
    print(n_main_paths)

    print("=" * 80)
    print("完成！仅输出以下内容：")
    print(f"1. {table_xlsx_path}")
    print(f"2. {n_main_paths}")
    print("=" * 80)


if __name__ == "__main__":
    main()