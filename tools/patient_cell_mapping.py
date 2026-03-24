'''
#有可解释性
# tools/patient_cell_mapping.py
import os
from collections import defaultdict
from pathlib import Path
import time
from tqdm import tqdm
import pandas as pd

def build_patient_image_map(root_dir):
    """构建患者到图像的映射"""
    print(f"🔍 扫描患者大图目录: {root_dir}")
    mapping = defaultdict(set)
    img_exts = {'.jpg', '.jpeg', '.png', '.bmp', '.webp'}
    
    patient_folders = [f for f in os.listdir(root_dir) if os.path.isdir(os.path.join(root_dir, f))]
    
    if not patient_folders:
        image_files = [f for f in os.listdir(root_dir) if Path(f).suffix.lower() in img_exts]
        for filename in image_files:
            name = Path(filename).stem
            mapping[name].add(name)
        return dict(mapping)

    for folder in tqdm(patient_folders, desc="Mapping Patients"):
        folder_path = os.path.join(root_dir, folder)
        files = [f for f in os.listdir(folder_path) if Path(f).suffix.lower() in img_exts]
        for f in files:
            mapping[folder].add(Path(f).stem)
            
    return dict(mapping)

def build_patient_cell_map(cell_root_dir, patient_image_map):
    """映射 单细胞 -> 患者"""
    print(f"🔍 扫描单细胞目录: {cell_root_dir}")
    image_to_patient = {}
    for pid, imgs in patient_image_map.items():
        for img in imgs:
            image_to_patient[img] = pid

    cell_files = [f for f in os.listdir(cell_root_dir) if f.lower().endswith('.jpg')]
    print(f"📸 找到 {len(cell_files)} 个细胞，开始匹配...")
    
    cell_patient_map = {}
    matched = 0
    
    for cell_file in tqdm(cell_files, desc="Matching"):
        stem = Path(cell_file).stem 
        parts = stem.split('_')
        
        if parts[-1].isdigit(): # 去掉增强后缀
            parts = parts[:-1]
        
        found_pid = None
        for k in range(len(parts), 0, -1):
            query = "_".join(parts[:k])
            if query in image_to_patient:
                found_pid = image_to_patient[query]
                break
        
        if found_pid:
            cell_patient_map[stem] = found_pid
            matched += 1
            
    print(f"✅ 匹配成功: {matched} / {len(cell_files)}")
    return cell_patient_map

def save_mapping(mapping, output_path):
    df = pd.DataFrame(list(mapping.items()), columns=['Cell_Filename', 'Patient_ID'])
    df.sort_values('Cell_Filename', inplace=True)
    df.to_excel(output_path, index=False)
    print(f"💾 映射表已保存: {output_path}")

def process(img_dir, cell_dir, out_path, tag):
    print(f"\n=== 处理 {tag} ===")
    p_map = build_patient_image_map(img_dir)
    c_map = build_patient_cell_map(cell_dir, p_map)
    if c_map:
        save_mapping(c_map, out_path)
    else:
        print("❌ 无匹配结果")

if __name__ == '__main__':
    # 原始大图路径
    train_orig = "/root/autodl-tmp/data/train_imgs_251127/"
    val_orig   = "/root/autodl-tmp/data/val_imgs_251127/"
    
    # [修改] 增强后的单细胞路径
    train_cell = "/root/autodl-tmp/projects/mwh/SingleCellProject/feature_o/aug_o_f/train/"
    val_cell   = "/root/autodl-tmp/projects/mwh/SingleCellProject/feature_o/aug_o_f/val/"
    
    # [修改] 映射表输出路径
    train_out = "/root/autodl-tmp/projects/mwh/SingleCellProject/feature_o/aug_o_f/Train_mapping.xlsx"
    val_out   = "/root/autodl-tmp/projects/mwh/SingleCellProject/feature_o/aug_o_f/Val_mapping.xlsx"

    process(train_orig, train_cell, train_out, "训练集")
    process(val_orig, val_cell, val_out, "验证集")

'''

#无可解释性
# tools/patient_cell_mapping.py
import os
from collections import defaultdict
from pathlib import Path
import time
from tqdm import tqdm

# ============================== 核心逻辑函数 ==============================

def build_patient_image_map(root_dir):
    """构建患者到图像的映射（基于有JSON标注的图像）"""
    print(f"🔍 正在从 {root_dir} 构建患者-图像映射...")
    mapping = defaultdict(set)
    img_exts = {'.jpg', '.jpeg', '.png', '.bmp', '.webp'}

    # 获取所有子文件夹（假设每个文件夹名就是一个患者 ID）
    patient_folders = [f for f in os.listdir(root_dir) if os.path.isdir(os.path.join(root_dir, f))]
    
    if not patient_folders:
        print("⚠️ 警告: 未发现子文件夹，将尝试平铺结构处理...")
        image_files = [f for f in os.listdir(root_dir) if Path(f).suffix.lower() in img_exts]
        for filename in image_files:
            name = Path(filename).stem
            mapping[name].add(name)
        return dict(mapping)

    print(f"📁 发现 {len(patient_folders)} 个患者文件夹")

    for folder in tqdm(patient_folders, desc="处理患者文件夹"):
        folder_path = os.path.join(root_dir, folder)
        image_files = [f for f in os.listdir(folder_path) if Path(f).suffix.lower() in img_exts]

        for filename in image_files:
            name = Path(filename).stem
            # 只要该图像存在对应的 JSON 标注，就认为该图像是有效的
            json_file = f"{name}.json"
            if os.path.exists(os.path.join(folder_path, json_file)):
                mapping[folder].add(name)

    total_images = sum(len(images) for images in mapping.values())
    print(f"✅ 映射完成: {len(mapping)} 个患者, 包含 {total_images} 张标注图像")
    return dict(mapping)


def build_patient_cell_map(cell_root_dir, patient_image_map):
    """根据文件名逻辑，将增强后的单细胞匹配回所属患者"""
    print(f"🧪 正在从 {cell_root_dir} 建立单细胞-患者关系...")
    
    cell_patient_map = {}
    
    # 构建反向查找字典: {image_stem: patient_id}
    image_to_patient = {}
    for patient, images in patient_image_map.items():
        for img_stem in images:
            image_to_patient[img_stem] = patient

    # 获取增强后的所有细胞图片
    cell_files = [f for f in os.listdir(cell_root_dir) if os.path.isfile(os.path.join(cell_root_dir, f))]
    print(f"📸 发现 {len(cell_files)} 个单细胞文件，开始匹配...")

    matched = 0
    unmatched = 0
    start_time = time.time()

    for i, cell_filename in enumerate(cell_files):
        # 进度打印
        if i % 2000 == 0 and i > 0:
            print(f"   进度: {i}/{len(cell_files)} ({i/len(cell_files)*100:.1f}%) | 已匹配: {matched}")

        # 解析文件名。例如: 03269_G14_015_N_3.jpg
        stem = Path(cell_filename).stem
        parts = stem.split('_')
        
        # 1. 如果有增强后缀 (如 _0, _1)，先去掉
        if parts[-1].isdigit() and len(parts) > 1:
            parts = parts[:-1]
        
        # 2. 尝试寻找原始大图名 (通过不断缩减前缀进行匹配)
        patient_found = None
        for j in range(len(parts), 0, -1):
            candidate_img = '_'.join(parts[:j])
            if candidate_img in image_to_patient:
                patient_found = image_to_patient[candidate_img]
                break
        
        if patient_found:
            cell_patient_map[stem] = patient_found
            matched += 1
        else:
            unmatched += 1

    print(f"✨ 匹配完成! 用时: {time.time()-start_time:.1f}s | 成功: {matched}, 失败: {unmatched}")
    return cell_patient_map


def save_mapping(mapping_dict, output_xlsx):
    """保存为 CSV 和 Excel"""
    output_path = Path(output_xlsx)
    csv_path = output_path.with_suffix('.csv')

    # 使用 Pandas 保存 CSV
    try:
        import pandas as pd
        df = pd.DataFrame(list(mapping_dict.items()), columns=['Cell_Filename', 'Patient_ID'])
        df.sort_values(by='Cell_Filename', inplace=True)
        df.to_csv(csv_path, index=False)
        print(f"💾 CSV 已保存: {csv_path}")
        
        # 尝试保存 Excel
        df.to_excel(output_path, index=False)
        print(f"💾 Excel 已保存: {output_path}")
    except Exception as e:
        print(f"⚠️ 保存时遇到错误: {e}")


def process_dataset(image_dir, cell_dir, output_file, name):
    print(f"\n{'='*20} 处理 {name} 数据集 {'='*20}")
    img_map = build_patient_image_map(image_dir)
    final_map = build_patient_cell_map(cell_dir, img_map)
    if final_map:
        save_mapping(final_map, output_file)
    else:
        print("❌ 未产生任何有效映射，请检查目录或文件名逻辑。")


# ============================== 主程序执行 ==============================

if __name__ == '__main__':
    # --- 训练集路径配置 ---
    train_orig = "/root/autodl-tmp/data/train_imgs_251127/"
    train_aug  = "/root/autodl-tmp/projects/mwh/SingleCellProject/cellseg_m/train/"
    train_out  = "/root/autodl-tmp/projects/mwh/SingleCellProject/cellseg_m/Train_mapping.xlsx"

    process_dataset(train_orig, train_aug, train_out, "训练集(Train)")

    # --- 验证集路径配置 ---
    val_orig = "/root/autodl-tmp/data/val_imgs_251127/"
    val_aug  = "/root/autodl-tmp/projects/mwh/SingleCellProject/cellseg_m/val/"
    val_out  = "/root/autodl-tmp/projects/mwh/SingleCellProject/cellseg_m/Val_mapping.xlsx"

    process_dataset(val_orig, val_aug, val_out, "验证集(Val)")

    print("\n🎉 所有数据集映射任务已完成！")
    

'''
# tools/patient_cell_mapping.py - 服务器版（完整映射患者-细胞关系）
import os
from collections import defaultdict
from pathlib import Path
import time
from tqdm import tqdm

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def build_patient_image_map(root_dir):
    """构建患者到图像的映射（基于有JSON标注的图像）"""
    print(f"从 {root_dir} 构建患者-图像映射")
    mapping = defaultdict(dict)
    img_exts = {'.jpg', '.jpeg', '.png', '.bmp', '.webp'}

    # 假设有患者子文件夹（如本地）
    patient_folders = [f for f in os.listdir(root_dir) if os.path.isdir(os.path.join(root_dir, f))]
    if not patient_folders:
        # 如果平铺，假设每个图像是一个'患者'，用图像stem作为患者ID
        print("警告: 无子文件夹，假设平铺结构，使用图像stem作为患者ID")
        image_files = [f for f in os.listdir(root_dir) if Path(f).suffix.lower() in img_exts]
        for filename in image_files:
            name = Path(filename).stem
            json_file = f"{name}.json"
            if os.path.exists(os.path.join(root_dir, json_file)):
                mapping[name] = {name}  # 患者ID = 图像stem
        return dict(mapping)

    print(f"发现 {len(patient_folders)} 个患者文件夹")

    for folder in tqdm(patient_folders, desc="处理患者文件夹"):
        folder_path = os.path.join(root_dir, folder)
        mapping[folder] = set()

        image_files = [f for f in os.listdir(folder_path)
                       if Path(f).suffix.lower() in img_exts]

        for filename in image_files:
            name = Path(filename).stem
            json_file = f"{name}.json"
            if os.path.exists(os.path.join(folder_path, json_file)):
                mapping[folder].add(name)

    total_images = sum(len(images) for images in mapping.values())
    print(f"患者-图像映射完成: {len(mapping)} 患者, {total_images} 标注图像")
    return dict(mapping)


def build_patient_cell_map(cell_root_dir, patient_image_map):
    """构建细胞到患者的映射，并统计每个患者细胞数"""
    print(f"从 {cell_root_dir} 构建患者-细胞映射")

    cell_patient_map = {}
    patient_cell_count = defaultdict(int)

    # 反向映射 {image_stem: patient}
    image_patient_map = {}
    for patient, images in patient_image_map.items():
        for image in images:
            image_patient_map[image] = patient

    # 获取所有细胞文件
    print("扫描细胞文件...")
    cell_files = [f for f in os.listdir(cell_root_dir) if os.path.isfile(os.path.join(cell_root_dir, f))]
    print(f"发现 {len(cell_files)} 个细胞文件，开始处理...")

    matched_count = unmatched_count = 0
    log_interval = 1000
    start_time = time.time()

    for i, cell_filename in enumerate(cell_files):
        if i % log_interval == 0 and i > 0:
            elapsed = time.time() - start_time
            rate = i / elapsed
            remaining = (len(cell_files) - i) / rate if rate > 0 else 0
            print(f"进度: {i}/{len(cell_files)} ({i/len(cell_files)*100:.1f}%), "
                  f"速率: {rate:.1f}/s, 已用: {elapsed:.1f}s, 剩余: {remaining:.1f}s, "
                  f"匹配: {matched_count}, 未匹配: {unmatched_count}")

        # 解析细胞文件名
        stem = Path(cell_filename).stem  # 去掉扩展名
        parts = stem.split('_')
        if len(parts) < 3:
            unmatched_count += 1
            continue

        # 候选图像stem（处理增强后的 _i）
        image_candidates = []
        
        # 如果有增强后缀 _i，去掉它
        if parts[-1].isdigit():  # 如 _0, _1 等
            parts = parts[:-1]

        # 策略1: 去掉最后3部分 (count_label)
        if len(parts) >= 4:
            image_candidates.append('_'.join(parts[:-3]))

        # 策略2: 去掉最后2部分
        if len(parts) >= 3:
            image_candidates.append('_'.join(parts[:-2]))

        # 策略3: 前3部分（常见模式）
        if len(parts) >= 3:
            image_candidates.append('_'.join(parts[:3]))

        # 策略4: 整个前缀直到数字
        for j in range(len(parts)-1, 1, -1):
            candidate = '_'.join(parts[:j])
            if candidate in image_patient_map:
                image_candidates.append(candidate)
                break

        # 去重并查找
        patient_found = None
        for candidate in set(image_candidates):
            if candidate in image_patient_map:
                patient_found = image_patient_map[candidate]
                break

        if patient_found:
            cell_patient_map[stem] = patient_found
            patient_cell_count[patient_found] += 1
            matched_count += 1
        else:
            unmatched_count += 1

    total_time = time.time() - start_time
    print(f"完成: {len(cell_files)} 文件，用时 {total_time:.1f}s")
    print(f"匹配: {matched_count}, 未匹配: {unmatched_count}, 率: {matched_count / len(cell_files) * 100 if len(cell_files) > 0 else 0:.1f}%")

    return dict(cell_patient_map), dict(patient_cell_count)


def save_to_excel(cell_patient_map, output_file):
    """保存为 Excel，如无 openpyxl 则自动保存 CSV"""
    output_path = Path(output_file)
    output_path = Path(output_file)
    csv_path = output_path.with_suffix('.csv')

    # 先强制保存 CSV（最稳）
    import pandas as pd
    df = pd.DataFrame(list(cell_patient_map.items()), columns=['Cell Filename', 'Patient ID'])
    df.to_csv(csv_path, index=False)
    print(f"CSV 已保存（最保险）: {csv_path}")

    # 再尝试保存 Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Cell-Patient Mapping"

        headers = ["Cell Filename", "Patient ID"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for row, (cell_file, patient) in enumerate(cell_patient_map.items(), 2):
            ws.cell(row, 1, cell_file)
            ws.cell(row, 2, patient)

        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = (max_len + 2) * 1.2

        ws.freeze_panes = "A2"
        wb.save(output_path)
        print(f"Excel 保存成功: {output_path}")
    except ImportError:
        print("openpyxl 未安装，已跳过 Excel 保存，仅保留 CSV")
    except Exception as e:
        print(f"Excel 保存失败: {e}，但 CSV 已成功保存")


def process_dataset(image_dir, cell_dir, output_xlsx, dataset_name):
    print(f"\n{'=' * 60}\n处理 {dataset_name} 数据集\n{'=' * 60}")
    start = time.time()

    patient_image_map = build_patient_image_map(image_dir)
    cell_patient_map, patient_cell_count = build_patient_cell_map(cell_dir, patient_image_map)

    if cell_patient_map:
        save_to_excel(cell_patient_map, output_xlsx)
    else:
        print("无映射，跳过Excel")

    print(f"{dataset_name} 处理完成，用时 {time.time() - start:.2f}s")


if __name__ == '__main__':
    # Train
    train_image_dir = "/root/autodl-tmp/projects/mwh/SingleCellProject/yolo/large_images/train_imgs"
    train_cell_dir = "/root/autodl-tmp/projects/mwh/SingleCellProject/yolo/single_cell_aug/Train"
    train_output_xlsx = "/root/autodl-tmp/projects/mwh/SingleCellProject/yolo/single_cell_aug/Train_mapping.xlsx"

    process_dataset(train_image_dir, train_cell_dir, train_output_xlsx, "Train")

    # Val
    val_image_dir = "/root/autodl-tmp/projects/mwh/SingleCellProject/yolo/large_images/val_imgs"
    val_cell_dir = "/root/autodl-tmp/projects/mwh/SingleCellProject/yolo/single_cell_aug/Valid"
    val_output_xlsx = "/root/autodl-tmp/projects/mwh/SingleCellProject/yolo/single_cell_aug/Valid_mapping.xlsx"

    process_dataset(val_image_dir, val_cell_dir, val_output_xlsx, "Val")

    print("\n所有处理完成！")

'''